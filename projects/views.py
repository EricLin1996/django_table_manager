from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from PIL import Image, ImageDraw, ImageFont
import io
import random
import string
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
from .models import SubProject, UserProfile
from .forms import LoginForm, SubProjectForm, CreateUserForm, UpdateUserForm


def generate_captcha_image(text):
    """生成验证码图片"""
    width, height = 120, 40
    image = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(image)
    
    # 尝试使用系统字体，如果没有则使用默认字体
    try:
        font = ImageFont.truetype("arial.ttf", 20)
    except:
        font = ImageFont.load_default()
    
    # 绘制验证码文字
    text_width = draw.textlength(text, font=font)
    text_height = 20
    x = (width - text_width) // 2
    y = (height - text_height) // 2
    
    # 添加一些随机色彩和干扰线
    for i in range(5):
        x1 = random.randint(0, width)
        y1 = random.randint(0, height)
        x2 = random.randint(0, width)
        y2 = random.randint(0, height)
        draw.line([(x1, y1), (x2, y2)], fill=(random.randint(200, 255), random.randint(200, 255), random.randint(200, 255)))
    
    draw.text((x, y), text, fill='black', font=font)
    
    # 将图片保存到字节流
    img_buffer = io.BytesIO()
    image.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer


def captcha_image(request):
    """生成验证码图片"""
    captcha_text = ''.join(random.choices(string.ascii_uppercase, k=4))
    request.session['captcha'] = captcha_text
    
    img_buffer = generate_captcha_image(captcha_text)
    
    response = HttpResponse(img_buffer.getvalue(), content_type='image/png')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response


def login_view(request):
    """登录视图"""
    if request.user.is_authenticated:
        return redirect('project_list')
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        
        # 手动验证验证码
        captcha_input = request.POST.get('captcha', '').upper()
        session_captcha = request.session.get('captcha', '').upper()
        
        if captcha_input != session_captcha:
            messages.error(request, '验证码错误')
            form = LoginForm()  # 重新生成表单和验证码
        elif form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'欢迎，{user.username}！')
            return redirect('project_list')
        else:
            messages.error(request, '登录失败，请检查用户名和密码')
    else:
        form = LoginForm()
    
    return render(request, 'projects/login.html', {'form': form})


def logout_view(request):
    """登出视图"""
    logout(request)
    messages.info(request, '您已成功登出')
    return redirect('login')


@login_required
def export_projects_excel(request):
    """导出项目到Excel（根据用户权限过滤）"""
    # 获取项目数据，根据用户权限过滤
    projects = SubProject.objects.all()
    
    # 检查用户权限并过滤数据
    try:
        profile = UserProfile.objects.get(user=request.user)
        is_admin = profile.is_admin
        user_organization = profile.organization
        
        # 如果不是管理员且有所属单位，则只导出该单位的项目
        if not is_admin and user_organization:
            projects = projects.filter(main_construction_unit=user_organization)
    except UserProfile.DoesNotExist:
        # 如果没有用户配置文件，默认只能导出空数据
        projects = SubProject.objects.none()
    
    projects = projects.order_by('-created_at')
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "子项目详细信息表"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='微软雅黑', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义表头
    headers = [
        '项目编号', '名称', '项目状态', '建设单位', '地址', '联系人', '电话',
        '面积(平方米)', '开工日期', '竣工日期', '施工单位', '施工单位联系人',
        '业主', '建安造价(元)', '日期', '工程期限', '巡查日期明细(次)',
        '现场开单记录明细(项)', '整改完成情况明细(项)', '备注', '创建时间', '更新时间'
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row, project in enumerate(projects, 2):
        data = [
            project.project_number or '',
            project.sub_project or '',
            project.get_status_display_custom() if project.status else '',
            project.construction_unit or '',
            project.address or '',
            project.contact_person or '',
            project.phone or '',
            float(project.area) if project.area else '',
            project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
            project.completion_date.strftime('%Y-%m-%d') if project.completion_date else '',
            project.contractor or '',
            project.contractor_contact or '',
            project.owner or '',
            float(project.construction_cost) if project.construction_cost else '',
            project.date.strftime('%Y-%m-%d') if project.date else '',
            project.construction_period or '',
            project.inspection_details if project.inspection_details is not None else '',
            project.on_site_records if project.on_site_records is not None else '',
            project.rectification_details if project.rectification_details is not None else '',
            project.remarks or '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            project.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
    
    # 调整列宽
    column_widths = [
        8, 12, 20, 25, 20, 30, 25, 15, 12, 12,
        12, 12, 30, 20, 10, 15, 20, 15, 15, 12,
        15, 30, 30, 20, 20
    ]
    
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 30  # 表头行高
    for row in range(2, len(projects) + 2):
        ws.row_dimensions[row].height = 25
    
    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # 生成文件名（支持中文）
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'子项目详细信息_{current_time}.xlsx'
    
    # 对文件名进行URL编码以支持中文
    encoded_filename = quote(filename.encode('utf-8'))
    
    # 设置Content-Disposition头，支持中文文件名
    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
    
    # 保存工作簿到响应
    wb.save(response)
    
    return response


@login_required
def project_list(request):
    """项目列表视图"""
    search_query = request.GET.get('search', '')
    projects = SubProject.objects.all()
    
    # 检查用户权限并过滤数据
    is_admin = False
    user_organization = None
    try:
        profile = UserProfile.objects.get(user=request.user)
        is_admin = profile.is_admin
        user_organization = profile.organization
        
        # 如果不是管理员且有所属单位，则只显示该单位的项目
        if not is_admin and user_organization:
            projects = projects.filter(main_construction_unit=user_organization)
    except UserProfile.DoesNotExist:
        # 如果没有用户配置文件，默认只能看到空数据
        projects = SubProject.objects.none()
    
    if search_query:
        projects = projects.filter(
            Q(project_number__icontains=search_query) |
            Q(sub_project__icontains=search_query) |
            Q(main_construction_unit__icontains=search_query) |
            Q(construction_unit__icontains=search_query) |
            Q(address__icontains=search_query)
        )
    
    # 获取总数量用于显示
    total_count = projects.count()
    
    paginator = Paginator(projects, 20)  # 每页显示20条记录
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'is_admin': is_admin,
        'total_count': total_count,
        'user_organization': user_organization,
    }
    
    return render(request, 'projects/project_list.html', context)


@login_required
def project_detail(request, pk):
    """项目详情视图"""
    project = get_object_or_404(SubProject, pk=pk)
    
    # 检查用户权限
    try:
        profile = UserProfile.objects.get(user=request.user)
        is_admin = profile.is_admin
        user_organization = profile.organization
        
        # 如果不是管理员且有所属单位，检查项目是否属于该单位
        if not is_admin and user_organization:
            if project.main_construction_unit != user_organization:
                messages.error(request, '您没有权限查看此项目')
                return redirect('project_list')
    except UserProfile.DoesNotExist:
        messages.error(request, '您没有权限查看此项目')
        return redirect('project_list')
    
    context = {
        'project': project,
    }
    
    return render(request, 'projects/project_detail.html', context)


@login_required
def project_edit(request, pk):
    """项目编辑视图"""
    project = get_object_or_404(SubProject, pk=pk)
    
    # 检查用户权限
    try:
        profile = UserProfile.objects.get(user=request.user)
        is_admin = profile.is_admin
        user_organization = profile.organization
        
        # 如果不是管理员且有所属单位，检查项目是否属于该单位
        if not is_admin and user_organization:
            if project.main_construction_unit != user_organization:
                messages.error(request, '您没有权限编辑此项目')
                return redirect('project_list')
    except UserProfile.DoesNotExist:
        messages.error(request, '您没有权限编辑此项目')
        return redirect('project_list')
    
    if request.method == 'POST':
        form = SubProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, '项目信息已成功更新')
            return redirect('project_detail', pk=project.pk)
        else:
            messages.error(request, '更新失败，请检查输入信息')
    else:
        form = SubProjectForm(instance=project)
    
    context = {
        'form': form,
        'project': project,
    }
    
    return render(request, 'projects/project_edit.html', context)


@login_required
def create_user(request):
    """创建用户视图（仅管理员）"""
    # 检查用户是否为管理员
    try:
        profile = UserProfile.objects.get(user=request.user)
        if not profile.is_admin:
            messages.error(request, '您没有权限创建用户')
            return redirect('project_list')
    except UserProfile.DoesNotExist:
        messages.error(request, '您没有权限创建用户')
        return redirect('project_list')
    
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'用户 {user.username} 创建成功')
            return redirect('user_list')
        else:
            messages.error(request, '创建用户失败，请检查输入信息')
    else:
        form = CreateUserForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'projects/create_user.html', context)


@login_required
def user_list(request):
    """用户列表视图（仅管理员）"""
    # 检查用户是否为管理员
    try:
        profile = UserProfile.objects.get(user=request.user)
        if not profile.is_admin:
            messages.error(request, '您没有权限查看用户列表')
            return redirect('project_list')
    except UserProfile.DoesNotExist:
        messages.error(request, '您没有权限查看用户列表')
        return redirect('project_list')
    
    profiles = UserProfile.objects.select_related('user').all()
    
    context = {
        'profiles': profiles,
    }
    
    return render(request, 'projects/user_list.html', context)


@login_required
def edit_user(request, user_id):
    """编辑用户视图（仅管理员）"""
    # 检查用户是否为管理员
    try:
        profile = UserProfile.objects.get(user=request.user)
        if not profile.is_admin:
            messages.error(request, '您没有权限编辑用户')
            return redirect('project_list')
    except UserProfile.DoesNotExist:
        messages.error(request, '您没有权限编辑用户')
        return redirect('project_list')
    
    edit_user = get_object_or_404(User, id=user_id)
    edit_profile = get_object_or_404(UserProfile, user=edit_user)
    
    if request.method == 'POST':
        form = UpdateUserForm(request.POST, instance=edit_user, user_profile=edit_profile)
        if form.is_valid():
            form.save()
            messages.success(request, f'用户 {edit_user.username} 信息已成功更新')
            return redirect('user_list')
        else:
            messages.error(request, '更新失败，请检查输入信息')
    else:
        form = UpdateUserForm(instance=edit_user, user_profile=edit_profile)
    
    context = {
        'form': form,
        'edit_user': edit_user,
        'edit_profile': edit_profile,
    }
    
    return render(request, 'projects/edit_user.html', context) 